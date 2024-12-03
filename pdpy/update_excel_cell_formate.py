import pathlib

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment
from collections import OrderedDict

file_path = '/Users/fq/Desktop/work/media/resource_excel_export/TBoard1668129096的副本.xlsx'
p = pathlib.Path(file_path)
df: pd.DataFrame = pd.read_excel(file_path)

computer_columns = ['续航时间', '前台任务时间', '待机时间', 'FC', 'ANR', 'UFT']

df = df.replace(-1, np.NaN)
# print(df.dtypes)
# median_dict = OrderedDict({'设备名称': '平均值'})
# for columns in computer_columns:
#     if columns in ['FC', 'ANR', 'UFT']:
#         median_dict[columns] = int(df[columns].median())
#     else:
#         median = int(df[columns].median())
#         median_dict[columns] = f"{median // 60 // 60}h{median // 60 % 60}min{median % 60}s"
#         for index, data in enumerate(df[columns]):
#             if not pd.isnull(data) and data:
#                 data = int(data)
#                 second = data % 60
#                 minute = data // 60 % 60
#                 hours = data // 60 // 60
#                 df.loc[index, columns] = f"{hours}h{minute}min{second}s"


# df = df.append(median_dict, ignore_index=True)
df.fillna('', inplace=True)

# for index, data in enumerate(df['前台任务时间']):
#     if data:
#         second = data % 60
#         minute = data // 60 % 60
#         hours = data // 60 // 60
#         df.loc[index, '前台任务时间'] = f"{hours}h{minute}min{second}s"
#         print(index, data)

print(f"shape: {df.shape}")


def make_hyperlink(value):
    url = "https://custom.url/{}"
    return '=HYPERLINK("%s", "%s")' % (value, value)
df['日志2'] = df['日志'].apply(lambda x: make_hyperlink(x))
print(df.to_string())

from openpyxl.styles import Font, builtins, NamedStyle

with pd.ExcelWriter('2.xlsx') as writer:
    df.to_excel(writer, engine='openpyxl', index=False)
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    reef_style = NamedStyle('reef_style')
    reef_style.alignment = Alignment(horizontal='center', vertical='center')
    # workbook.add_named_style(reef_style)
    # column add style
    # for column in range(1, worksheet.max_column + 1):
    #     col_char = openpyxl.utils.get_column_letter(column)
    #     col = worksheet.column_dimensions[col_char]
    #     col.alignment = Alignment(horizontal='center', vertical='center')

    for index in range(1, worksheet.max_row + 1):
        row = worksheet.row_dimensions[index]
        row.alignment = Alignment(horizontal='center', vertical='center')
    # worksheet.style = reef_style
    # for cell in worksheet['J']:
    #     cell.alignment = Alignment(horizontal='center', wrapText=True)
    #     hyperlink = Font(underline='single', color='0072E3')
    #     cell.font = hyperlink
    #     link = cell.hyperlink
    #     print(f'link: {link}')

