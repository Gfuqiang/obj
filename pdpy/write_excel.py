import openpyxl
import pandas as pd
import pathlib
# dict_data = {'a': [1, 2, 3], 'b': [2,3]}
# columns = list(dict_data.keys())
# data = list(dict_data.values())
# print(list(zip(columns, data)))

# 以等号开头在excel中消失问题
from openpyxl.styles import Alignment, Font

df = pd.DataFrame(["=5", "=3", "=4"])
print(df)

p = pathlib.Path('1.xlsx')

with pd.ExcelWriter(p, engine='openpyxl') as writer:

# Convert the dataframe to an XlsxWriter Excel object.


# Get the xlsxwriter objects from the dataframe writer object.
    workbook = writer.book
    # worksheet = workbook.active
    worksheet = writer.sheets['Sheet1']
    df.to_excel(writer, sheet_name='Sheet1', index=False)
# worksheet.cell(2, 2).number_format = '@'
    for index, columns in enumerate(df.columns):
        print(columns)
        print(f"word：{openpyxl.utils.get_column_letter(index + 1)}")
    cola = worksheet.column_dimensions['A']
    for cell in worksheet['A']:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.data_type == 'f':
            cell.data_type = 's'


# worksheet['A3'].data_type = 's'
# worksheet['A2'].data_type = 's'


# worksheet.cell(2, 2).alignment = Alignment(horizontal="center", vertical="center")
# worksheet.cell(2, 2).data_type = 's'

# print(col.number_format)


