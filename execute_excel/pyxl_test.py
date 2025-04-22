import datetime
import os
import pathlib

import openpyxl
import pandas as pd
from openpyxl import load_workbook

def openpyxl_test():
    # 加载原始Excel文件
    wb = load_workbook('/Users/fq/Desktop/work/obj/execute_excel/年份测试.xlsx', data_only=True)

    # 选择要操作的工作表
    ws = wb.active

    # 修改单元格内容
    # ws['A1'] = 'Hello, World!'
    print(f'{ws}')
    # 保存修改后的Excel文件（保留样式）
    wb.save(f'/Users/fq/Desktop/work/obj/execute_excel/年份测试{datetime.datetime.now()}.xlsx')

    print("Excel文件已成功修改并保存，样式已保留。")

def value_error_test():
    file_path = '/Users/fq/Desktop/work/obj/execute_excel/年份测试.xlsx'
    df = pd.read_excel(file_path)
    df = df.fillna('')
    wb = load_workbook(file_path)
    ws = wb.active
    new_file_path = f'/Users/fq/Desktop/work/obj/execute_excel/年份测试{datetime.datetime.now()}.xlsx'
    for index, row in df.iterrows():
        line = int(index) + 2
        total = row['代收金额']
        income = row['到账收入']
        if income == '':
            df.loc[index, '到账收入'] = 0.00
        if total == '':
            print(f'{line}')
            df.loc[index, '代收金额'] = 0.00

    print(df.iloc[2])
    print(df.dtypes)
    with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
        # 加载原有的工作簿
        writer.book = wb
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        # 写入修改后的 DataFrame
        df.to_excel(writer, index=False)

    # wb.save(new_file_path)

def solve_openpyxl_question():
    # 读取Excel文件
    workbook = openpyxl.load_workbook('/Users/fq/Desktop/work/obj/execute_excel/年份测试.xlsx')
    sheet = workbook['工作表1']

    # 查看公式
    for row in sheet.iter_rows(values_only=True):
        for cell in row:
            if cell.data_type == 'f':
                print(cell.value)

if __name__ == '__main__':
    # value_error_test()
    # openpyxl_test()
    # solve_openpyxl_question()
    file_path = '/Users/fq/Desktop/work/obj/execute_excel/年份测试.xlsx'
    path = pathlib.Path(file_path)
    print(path.suffix)
    print(path.name)
    print(path.parent)
    file_name = path.name.split(".")[0] + f'_{datetime.datetime.now()}' + path.suffix
    # new_file_path = os.path.join(path.parent, file_name)
    new_file_path = path.parent / file_name
    print(new_file_path)
    # file_name = path.parents[0] + path.name
    # print(file_name)
