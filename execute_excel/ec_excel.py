"""
log error 和 waring 分开写
"""
import datetime
import pathlib

import numpy as np
import pandas as pd
from PySide6.QtUiTools import QUiLoader
from PySide6.QtWidgets import QApplication, QMainWindow, QMessageBox, QFileDialog
from openpyxl import load_workbook

class LoaderQT:

    def __init__(self):
        self.ui = QUiLoader().load("ec_excel.ui")
        self.ui.upload_file.clicked.connect(self.upload_file)
        self.ui.date_check.clicked.connect(self.date_check)

    def date_check(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self.ui, "Select excel file", '', 'Excel files (*.xls *.xlsx *.py)')
        self.write_logs(f'-' * 100)
        self.write_logs(f'文件路径：{file_path}')
        if not file_path:
            self.write_logs(f'未发现选中的文件，程序停止执行')

        handle_excel = HandleExcel(file_path, loaderqt=self)
        handle_excel.date_check()

    def upload_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self.ui, "Select excel file", '', 'Excel files (*.xls *.xlsx *.py)')
        self.write_logs(f'程序使用说明。程序读取 excel 文件的每行数据，根据以下规则判断是否执行计算逻辑：\n'
                        f'1.公户到账金额、出款金额都有数据且 为出款金额项是 “已回公户未出”的要算。\n'
                        f'2.公户到账金额、未出款金额列都有数据且不相同\n'
                        f'需要计算的列及计算规则：\n'
                        f'1.公户未出金额：公户到账金额-对私到账金额。\n'
                        f'2.到账收入：合同金额 *（报价-成本)。\n'
                        f'3.代收金额：合同金额 * 成本。\n'
                        f'4.应回款金额：公户到账金额 - 到账收入 - 代收金额。\n'
                        )
        self.write_logs(f'-' * 100)
        self.write_logs(f'文件路径：{file_path}')
        if not file_path:
            self.write_logs(f'未发现选中的文件，程序停止执行')

        # handle_excel = HandleExcel('/Users/fq/Desktop/work/obj/execute_excel/年份测试.xlsx')
        handle_excel = HandleExcel(file_path, loaderqt=self)
        handle_excel.execute_data()


    def write_logs(self, text:str):
        self.ui.logs.append(text)
        self.ui.logs.ensureCursorVisible()


def parameter_handle(func):
    def wrapper(self, *args, **kwargs):
        new_kwargs = {}
        for key, value in kwargs.items():
            if isinstance(value, float):
                new_kwargs[key] = value
                continue
            new_kwargs[key] = transform_int(value)
        return func(self, *args, **new_kwargs)
    return wrapper

def transform_int(input):
    try:
        input = float(input)
    except ValueError:
        pass
    return input


class HandleExcel:

    def __init__(self, file_path:str, loaderqt: LoaderQT|None=None ):
        self.file_path = file_path
        self.loader_qt = loaderqt
        self.update_row_list = []
        self.new_file_path = self.new_file_path()

    def new_file_path(self):
        path = pathlib.Path(self.file_path)
        file_name = path.name.split(".")[0] + f'_{datetime.datetime.now()}' + path.suffix
        # new_file_path = os.path.join(path.parent, file_name)
        new_file_path = path.parent / file_name
        return new_file_path

    def print_log(self, text:str):
        if self.loader_qt is not None:
            self.loader_qt.write_logs(text)
        else:
            print(text)

    def error_log(self, text:str, line):
        self.print_log(f'[错误] 行数 {line}:\n  ' + text)

    def info_log(self, text:str, line):
        self.print_log(f'[计算成功] 行数 {line}:\n    ' + text)

    def waring_log(self, text:str, line):
        self.print_log(f'[更新错误提示] 行数 {line}:\n  ' + text)

    def date_info_log(self, text:str, line):
        self.print_log(f'[回款提示] 行数 {line}:\n    ' + text)

    def read_excel(self):
        path = pathlib.Path(self.file_path)
        df = pd.read_excel(path)
        # print(df)
        return df

    def date_check(self):
        df = self.read_excel()
        df = df.fillna('')
        for index, row in df.iterrows():
            line = index + 2
            partyA = row['甲方打款日期']
            partyB = row['乙方回款日期（垫款日期）']
            if not partyA:
                self.date_info_log(f'缺少"甲方打款日期"数据', line)
                continue
            try:
                if isinstance(partyA, str):
                    print(f'{partyA}')
                    if '、' in partyA:
                        partyA = partyA.split('、')[0]
                    partyA = datetime.datetime.strptime(partyA, '%Y/%m/%d')
                end_day = partyA + datetime.timedelta(days=4)
                now_day = datetime.datetime.now()
                if now_day > end_day and partyB == '':
                    self.date_info_log(f'回款日期已逾期。 甲方打款日期是：{partyA}, 当前日期是：{now_day}', line)
            except ValueError:
                self.error_log(f'甲方打款日期格式 {partyA} 有误不能转换格式需要为 "2025/x/xx" 或 "2025/x/xx、x/xx"', line)

    def execute_data(self):
        self.print_log(f'开始读取 excel 文件')
        df = self.read_excel()
        self.print_log(f'excel 文件读取完成')
        df = df.fillna('')
        compute_count = 0
        for index, row in df.iterrows():
            old_index = index
            line = index + 2
            public_account_amount = row['公户到账金额']
            payment_amount = row['出款金额']
            no_payment_amount = row['未出款金额']
            private_account_arrived_amount = row['对私到账金额']
            contract_amount = row['合同金额']
            quotation = row['报价']
            cost = row['成本']
            revenue_received = row['到账收入']

            # if public_account_amount != '' and payment_amount != '' and payment_amount != public_account_amount:
            #     # 用于提示
            #     print(f'index: {index}, 公户到账金额和出款金额不相同')
                # self.risk_warning('公户到账金额和出款金额不相同')

            if public_account_amount == payment_amount and no_payment_amount != '已回公户未出':
                # 公户到账金额 = 出款金额 不计算
                continue

            if self.compute_judge(row, line):
                # 符合条件进行计算
                self.print_log(f'-' * 100)
                self.print_log(f'第 {line} 行符合计算规则开始计算:')
                try:
                    ret_public_account_no_payment_amount = self.public_account_no_payment_amount(
                        public_account_amount=public_account_amount,
                        private_account_arrived_amount=private_account_arrived_amount,
                    )
                except Exception as e:
                    self.error_log(
                        f'[公户未出款金额] 计算错误，公户到账金额：{public_account_amount},'
                        f' 对私到账金额: {private_account_arrived_amount}', line)
                    continue

                try:
                    ret_income_received = self.income_received(
                        contract_amount=contract_amount,
                        quotation=quotation,
                        cost=cost,
                    )
                except Exception as e:
                    self.error_log(
                        f'[到账收入] 计算错误，合同金额：{contract_amount},'
                        f'报价 : {quotation}, 成本：{cost}', line)
                    continue

                ret_collection_amount = self.collection_amount(
                    contract_amount=contract_amount,
                    cost=cost,
                )
                try:
                    ret_amount_receivable = self.amount_receivable(
                        public_account_amount=public_account_amount,
                        revenue_received=revenue_received,
                        collection_amount=ret_collection_amount
                    )
                except Exception as e:
                    self.error_log(
                        f'[应回款金额]计算错误，合同金额：{contract_amount},'
                        f'公户到账金额 : {public_account_amount},'
                        f' 到账收入：{revenue_received}, 代收金额：{ret_collection_amount}', line)
                    continue

                self.info_log(
                    f' 合同金额：{int(contract_amount)}, 公户未出款金额:{ret_public_account_no_payment_amount},'
                    f' 到账收入:{ret_income_received}, 代收金额:{ret_collection_amount},'
                    f' 应回款金额:{ret_amount_receivable}', line
                )

                # change column
                update_data = {
                    '公户未出金额': ret_public_account_no_payment_amount,
                    '到账收入': ret_income_received,
                    '代收金额': ret_collection_amount,
                    '应回款金额': ret_amount_receivable,
                }
                update_line_data = True
                for change_column, new_val in update_data.items():
                    update_ret = self.update_df(df, row, line, old_index, change_column, new_val)
                    if not update_ret:
                        update_line_data = False
                if update_line_data:
                    compute_count += 1
                    self.update_row_list.append(line)
        self.print_log(f'更新数据 {compute_count} 条\n更新的行数有：{self.update_row_list}')
        df.replace('', np.nan, inplace=True)
        # print(f"合计：{df.iloc[17]}")
        # print(f"类型：{df.dtypes}")
        self.update_sources_file(df)

    def compute_judge(self, row, index):
        public_account_amount = row['公户到账金额']
        payment_amount = row['出款金额']
        no_payment_amount = row['未出款金额']

        if public_account_amount != '' and payment_amount != '' and no_payment_amount == '已回公户未出':
            return True
        if public_account_amount != '' and no_payment_amount:
            return True
        return False

    def update_df(self, df, row, line, old_index, change_column, new_value):
        """
        待更新列中已有数据的，所有列都不会进行更新
        :param df:
        :param row:
        :param line:
        :param old_index:
        :param change_column:
        :param new_value:
        :return:
        """
        if row[change_column] != '':
            self.waring_log(f'{change_column}列已有数据，程序不进行修改请自行校验。', line)
            return
        df.loc[old_index, change_column] = new_value
        return True

    def update_sources_file(self, df):
        # df.to_excel(self.file_path, index=False)
        # 将修改后的 DataFrame 写回原文件，同时保留样式
        wb = load_workbook(self.file_path, data_only=True)
        ws = wb.active
        with pd.ExcelWriter(self.new_file_path, engine='openpyxl') as writer:
            # 加载原有的工作簿
            writer.book = wb
            # writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
            # 写入修改后的 DataFrame
            df.to_excel(writer, index=False, startrow=ws.max_row + 2)  # 替换为你的工作表名称
        print('*' * 20 + '文件写入完成' + '*' * 20)

    @parameter_handle
    def public_account_no_payment_amount(self, public_account_amount, private_account_arrived_amount):
        """
        公户未出金额：公户到账金额-对私到账金额
        :return:
        """
        return round(public_account_amount - private_account_arrived_amount, 2)

    @parameter_handle
    def income_received(self, contract_amount, quotation, cost):
        """
        到账收入：合同金额 *（报账-成本）
        :return:
        """
        return round(contract_amount * (quotation - cost), 2)

    @parameter_handle
    def collection_amount(self, contract_amount, cost):
        """
        代收金额：合同金额 * 成本
        :return:
        """
        return round(contract_amount * cost, 2)

    @parameter_handle
    def amount_receivable(self, public_account_amount, revenue_received, collection_amount):
        """
        应回款金额：公户到账金额 - 到账收入 - 代收金额
        :return:
        """
        return round(public_account_amount - revenue_received - collection_amount, 2)

    def risk_warning(self, text):
        self.loader_qt.write_logs(f'==============={text}===============')

if __name__ == '__main__':
    app = QApplication([])
    load_qt = LoaderQT()
    load_qt.ui.show()
    app.exec()

    # 单独测试 excel 功能
    # handle_excel = HandleExcel('/Users/fq/Desktop/work/obj/execute_excel/年份测试.xlsx')
    # handle_excel.execute_data()
